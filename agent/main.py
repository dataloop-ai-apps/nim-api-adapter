"""
Compare NVIDIA NIM models with Dataloop Marketplace.
Simple standalone script.
"""

import os
from datetime import datetime
from openai import OpenAI
import dtlpy as dl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def fetch_nvidia_nim_models():
    """Fetch all available NIM models from NVIDIA via OpenAI-compatible API."""
    print("\n🔍 Fetching NIM models from NVIDIA...")
    
    api_key = os.environ.get("NVIDIA_API_KEY") or os.environ.get("NGC_API_KEY")
    if not api_key:
        print("❌ NVIDIA_API_KEY or NGC_API_KEY not set")
        return []
    
    try:
        
        client = OpenAI(
            base_url="https://integrate.api.nvidia.com/v1",
            api_key=api_key
        )
        
        response = client.models.list()
        models = [model.id for model in response.data]
        
        print(f"✅ Found {len(models)} NIM models")
        return models
    except Exception as e:
        print(f"❌ Error fetching models: {e}")
        return []


def fetch_dataloop_nim_dpks():
    """Fetch all public DPKs from Dataloop with NIM category."""
    print("\n🔍 Fetching NIM DPKs from Dataloop...")
    
    filters = dl.Filters(resource=dl.FiltersResource.DPK)
    filters.add(field='scope', values='public')
    filters.add(field='attributes.Category', values='NIM')
    
    dpks = dl.dpks.list(filters=filters)
    
    nim_dpks = []
    for dpk in dpks.all():
        nim_dpks.append({
            "name": dpk.name,
            "display_name": dpk.display_name,
            "version": dpk.version
        })
    
    print(f"✅ Found {len(nim_dpks)} NIM DPKs in Dataloop")
    return nim_dpks


def normalize_name(name: str) -> str:
    """Normalize model name for comparison."""
    return name.lower().replace("/", "-").replace("_", "-").replace(" ", "-")


def compare_models(nim_models: list, dataloop_dpks: list) -> dict:
    """Compare NIM models with Dataloop DPKs."""
    
    # Normalize names for comparison
    nim_normalized = {normalize_name(m): m for m in nim_models}
    dataloop_normalized = {normalize_name(d["name"]): d for d in dataloop_dpks}
    
    # Find missing (in NIM but not in Dataloop)
    missing = []
    matched_nim = []
    
    for nim_norm, nim_original in nim_normalized.items():
        found = False
        for dl_norm in dataloop_normalized.keys():
            if nim_norm in dl_norm or dl_norm in nim_norm:
                found = True
                matched_nim.append(nim_original)
                break
        if not found:
            missing.append(nim_original)
    
    # Find deprecated (in Dataloop but not in NIM)
    deprecated = []
    matched_dl = []
    
    for dl_norm, dl_dpk in dataloop_normalized.items():
        found = False
        for nim_norm in nim_normalized.keys():
            if nim_norm in dl_norm or dl_norm in nim_norm:
                found = True
                matched_dl.append(dl_dpk)
                break
        if not found:
            deprecated.append(dl_dpk)
    
    return {
        "missing": missing,          # In NIM, not in Dataloop (need to add)
        "deprecated": deprecated,    # In Dataloop, not in NIM (potentially remove)
        "matched_nim": matched_nim,
        "matched_dl": matched_dl,
    }


def style_header(ws, row=1):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def auto_width(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def write_to_excel(nim_models: list, dataloop_dpks: list, comparison: dict):
    """Write all results to a single Excel file with multiple sheets."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"nim_comparison_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["NVIDIA NIM Models (Total)", len(nim_models)])
    ws_summary.append(["Dataloop NIM DPKs (Total)", len(dataloop_dpks)])
    ws_summary.append(["Matched (In Both)", len(comparison["matched_nim"])])
    ws_summary.append(["Missing in Dataloop (To Add)", len(comparison["missing"])])
    ws_summary.append(["Deprecated (To Remove)", len(comparison["deprecated"])])
    ws_summary.append([])
    ws_summary.append(["Generated", datetime.now().isoformat()])
    style_header(ws_summary)
    auto_width(ws_summary)
    
    # Sheet 2: All NIM Models
    ws_nim = wb.create_sheet("NIM Models")
    ws_nim.append(["#", "Model ID", "Publisher", "Status"])
    for i, model in enumerate(nim_models, 1):
        publisher = model.split("/")[0] if "/" in model else "nvidia"
        status = "✅ In Dataloop" if model in comparison["matched_nim"] else "❌ Missing"
        ws_nim.append([i, model, publisher, status])
    style_header(ws_nim)
    auto_width(ws_nim)
    
    # Sheet 3: Dataloop DPKs
    ws_dl = wb.create_sheet("Dataloop DPKs")
    ws_dl.append(["#", "DPK Name", "Display Name", "Version", "Status"])
    for i, dpk in enumerate(dataloop_dpks, 1):
        status = "✅ Active" if dpk in comparison["matched_dl"] else "⚠️ Deprecated"
        ws_dl.append([i, dpk["name"], dpk["display_name"], dpk["version"], status])
    style_header(ws_dl)
    auto_width(ws_dl)
    
    # Sheet 4: Missing (To Add)
    ws_missing = wb.create_sheet("To Add")
    ws_missing.append(["#", "Model ID", "Publisher", "Action"])
    for i, model in enumerate(comparison["missing"], 1):
        publisher = model.split("/")[0] if "/" in model else "nvidia"
        ws_missing.append([i, model, publisher, "ADD"])
    style_header(ws_missing)
    auto_width(ws_missing)
    # Color the sheet tab green
    ws_missing.sheet_properties.tabColor = "00B050"
    
    # Sheet 5: Deprecated (To Remove)
    ws_deprecated = wb.create_sheet("To Remove")
    ws_deprecated.append(["#", "DPK Name", "Display Name", "Version", "Action"])
    for i, dpk in enumerate(comparison["deprecated"], 1):
        ws_deprecated.append([i, dpk["name"], dpk["display_name"], dpk["version"], "REMOVE"])
    style_header(ws_deprecated)
    auto_width(ws_deprecated)
    # Color the sheet tab red
    ws_deprecated.sheet_properties.tabColor = "FF0000"
    
    # Save
    wb.save(filename)
    print(f"📄 Written: {filename}")
    
    return filename


def main():
    print("=" * 60)
    print("NIM Models vs Dataloop Marketplace Comparison")
    print("=" * 60)
    
    # Fetch from NVIDIA
    nim_models = fetch_nvidia_nim_models()
    
    # Fetch from Dataloop
    dataloop_dpks = fetch_dataloop_nim_dpks()
    
    # Compare
    print("\n🔍 Comparing...")
    comparison = compare_models(nim_models, dataloop_dpks)
    
    # Display NIM models
    print("\n" + "=" * 60)
    print(f"NVIDIA NIM Models ({len(nim_models)})")
    print("=" * 60)
    for model in nim_models[:20]:
        status = "✅" if model in comparison["matched_nim"] else "❌"
        print(f"  {status} {model}")
    if len(nim_models) > 20:
        print(f"  ... and {len(nim_models) - 20} more")
    
    # Display Dataloop DPKs
    print("\n" + "=" * 60)
    print(f"Dataloop NIM DPKs ({len(dataloop_dpks)})")
    print("=" * 60)
    for dpk in dataloop_dpks:
        status = "✅" if dpk in comparison["matched_dl"] else "⚠️"
        print(f"  {status} {dpk['name']} (v{dpk['version']})")
    
    # Missing in Dataloop
    print("\n" + "=" * 60)
    print(f"🆕 Missing in Dataloop - TO ADD ({len(comparison['missing'])})")
    print("=" * 60)
    for model in comparison["missing"][:20]:
        print(f"  ➕ {model}")
    if len(comparison["missing"]) > 20:
        print(f"  ... and {len(comparison['missing']) - 20} more")
    
    # Deprecated in Dataloop
    print("\n" + "=" * 60)
    print(f"⚠️  Deprecated - TO REMOVE ({len(comparison['deprecated'])})")
    print("=" * 60)
    for dpk in comparison["deprecated"]:
        print(f"  🗑️  {dpk['name']} (v{dpk['version']})")
    if not comparison["deprecated"]:
        print("  (none)")
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 Summary")
    print("=" * 60)
    print(f"  NVIDIA NIM models:     {len(nim_models)}")
    print(f"  Dataloop NIM DPKs:     {len(dataloop_dpks)}")
    print(f"  Matched (in both):     {len(comparison['matched_nim'])}")
    print(f"  Missing (to add):      {len(comparison['missing'])}")
    print(f"  Deprecated (to remove):{len(comparison['deprecated'])}")
    
    # Write to Excel
    print("\n" + "=" * 60)
    print("📁 Writing Excel file...")
    print("=" * 60)
    filename = write_to_excel(nim_models, dataloop_dpks, comparison)
    
    print("\n✅ Done!")


if __name__ == "__main__":
    main()
